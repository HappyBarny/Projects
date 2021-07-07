# -*- coding: utf-8 -*-
"""
Created on Fri Feb 26 10:22:04 2021

@author: paolo
"""

'''
SEZIONE IN CUI AVVIENE IL WEB SCRAPING
'''

from bs4 import BeautifulSoup
import requests
from datetime import datetime

numberPages = None

def web_scraping(url):
    # url = "https://www.newegg.com/p/pl?d=graphic+card"
    page = requests.get(url)
    
    soup = BeautifulSoup(page.content, 'html.parser')
    
    '''
    WHERE EVERYTHING BEGINS: SCRAPING DEGLI ELEMENTI
    '''
    
    all_items = soup.findAll("div", {"class" : "item-cell"})
    oggetti = []
    
    for i in all_items:
        
        '''
        NAME SCRAPING
        '''
        tag_name = ('div.item-info a.item-title')
        if i.select_one(tag_name) != None:
            name = (i.select_one(tag_name).string)
            
            '''
            SHIPPING COST SCRAPING
            '''
            list_items = i.find_all('li') 
            list_contents = [e.string for e in list_items]
            for item in list_contents: #shipping cost
                if (item != None):
                    if (item !='\xa0'):
                        # print(item)
                        shipping_cost = item
                    
            '''
            PRICE SCRAPING
            '''
            tag_price_strong = 'li.price-current strong'
            tag_price_sup = 'li.price-current sup'
            cost_strong = (i.select_one(tag_price_strong))
            cost_sup = (i.select_one(tag_price_sup))
            if cost_strong != None:
                cost = cost_strong.string + cost_sup.string
            if cost_strong == None:
                cost = ""

            '''
            SIGN TIME
            ''' 
            data = datetime.now().strftime('%d-%m-%Y %H:%M')
            
            '''
            TAKING LINKS FOR EACH PRODUCT 
            '''
            link = i.a['href']
            
            
            '''
            MAKING A MATRIX WITH DATA
            '''
            oggetti.append({"name": name , "cost of the product" : cost, "shipping cost" : shipping_cost, "data" : data, "link" : link})
    
    
    return oggetti ##É una lista

#TROVA QUANTE PAGINE BISOGNA ITERARE
def page_numbers(url):
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    
    locator = 'span.list-tool-pagination-text strong'
    item = str(soup.select_one(locator))
    pages = int(item.replace('<strong>', '').replace('<!-- -->', '').replace('</strong>', '').replace('1/', ""))
    return pages


'''
WRITING DATA IN A JSON FILE
WARNING: IT DOESNT TAKE A TRACE FOR PRODUCTS IN LONG PERIOD DISTANCE. FILE NEED TO BE RESETTED EVERY TIME.
'''

import os
import json 

items_file = 'data.json'

def create_item_table():
    #controlla se il file è vuoto o meno, se è vuoto ne crea uno in formato json (il file creato è vuoto [])
    if os.stat(items_file).st_size == 0:    
        with open(items_file, 'w') as file:
            json.dump([], file)  
            
def reset_file_json():
    with open(items_file, 'w') as file:
        json.dump([], file)  
        print('file is resetted')

def add_items_file_json(item):
    itemlist = get_all_items_from_file_json()
    itemlist.append(item)
    _save_all_items_json(itemlist)

def _save_all_items_json(item): #private function
    with open(items_file, 'w') as file:
        json.dump(item, file)

def get_all_items_from_file_json():
    with open(items_file, 'r') as file:
        return json.load(file)
        
def get_specific_item_json(name):
    items_file = get_all_items_from_file_json()
    for item in items_file:
        if(name in item['name']):
            print(item['name'] + "\nPrice: " + item['cost of the product'] + " $" + '\nOther costs: '+ item['shipping cost'] + "\nLINK: " + item['link'] + "\n")

def search_product_json():
    name= input('ricerca prodotto: ')
    name = (name.upper())
    get_specific_item_json(name)    

'''
ONLY WRITING AND READING DATA IN A CSV FILE
'''
from openpyxl import Workbook
from openpyxl import load_workbook

workbook = Workbook()
filename = ""

def create_file(oggetti):
    print('new file is being created')
    sheet = workbook.active

    sheet['A1'] = 'NAME'    
    sheet['B1'] = 'COST'
    sheet["C1"] = 'SHIPPING COST'
    sheet["D1"] = 'DATA'
    sheet["E1"] = 'LINK'
        
    n = 2
    for i in oggetti:    
        cell = str(n)
        sheet["A"+cell] = i['name']
        sheet["B"+cell] = i['cost of the product']
        sheet["C"+cell] = i['shipping cost']
        sheet["D"+cell] = i['data']
        sheet["E"+cell] = i['link']
        n+=1
    
    global filename    
    workbook.save(filename)


def _read_file():
    global filename
    workbook = load_workbook(filename)
    sheet = workbook.active
    
    column = len(sheet['A'])+1
    listItems = []
    for i in range(1, column, 1):
        name = (sheet['A'+str(i)].value)
        cost = (sheet['B'+str(i)].value)
        shipping = (sheet['C'+str(i)].value)
        data = (sheet['D'+str(i)].value)
        link = (sheet['E'+str(i)].value)
        dic = {'name' : name, 'cost of the product' : cost, 'shipping cost' : shipping, 'data' : data , "link" : link}
        listItems.append(dic)
    return(listItems)            
        
def update_file(oggetti):
    listItems = _read_file()
    
    #obj lista di oggetti su cui è stato effettuato il web scraping
    #item lista di oggetti letti nel file
    for obj in oggetti:
        n = 0
        for item in listItems:
            if obj['name'] in item['name']:
                if obj['cost of the product'] != item['cost of the product']:
                    # print(item['name'] + ': price is changed. Current price: ' + item['cost of the product'] + '\n')
                    item['cost of the product'] = obj['cost of the product']
                    # print('VALUE IS CHANGED: ' + item['cost of the product'])
                
                if (obj['shipping cost'] != item['shipping cost']):
                    # print(item['name'] + ': shipping is changed. Current price: ' + item['shipping cost'] + '\n')
                    item['shipping cost'] = obj['shipping cost']
                    # print('VALUE IS CHANGED: ' + item['shipping cost'])
                
                item['data'] = obj['data']
                n = 1
                
        if n != 1:
            # print('object doesnt exists. Adding to list')
            listItems.append(obj)
            
    sheet = workbook.active        
    n = 1
    for item in listItems:
        cell = str(n)
        sheet["A"+cell] = item['name']
        sheet["B"+cell] = item['cost of the product']
        sheet["C"+cell] = item['shipping cost']
        sheet["D"+cell] = item['data']
        sheet["E"+cell] = item['link']
        n+=1
    
    global filename    
    workbook.save(filename)   


'''
CONSOLE INTERFACE
'''
#SOME KEY WORD FOR SOME PRODUCTS: GRAPHIC CARD, MONITOR
def give_url():
    
    parameter = input('please type the object that you wanna scrape: ').replace(' ', '+')
    
    # parameter = "graphic+card" #standard parameter use it only for test
    
    url = "https://www.newegg.com/p/pl?d=" + parameter    
    global filename 
    filename = "New_Egg_" + parameter.replace("+", "_") + ".xlsx"
    
    return url

def update_url(url, i):
    sUrl = url + "%3D2&page=" + str(i)    
    print(sUrl)
    
    return sUrl
    
    
    
'''
ESECUZIONE DEL PROGRAMMA
'''
import os.path
from os import path

url = give_url() #metodo della console interface

pageNumber = int(page_numbers(url)//20) #use this just for rapid checks
# pageNumber = int(page_numbers(url))

for i in range(1, pageNumber, 1):
    sUrl = update_url(url, i)
    oggetti = web_scraping(sUrl)
    
# oggetti = web_scraping('https://www.newegg.com/p/pl?d=graphic+card%3D2&page=36')

create_item_table() #crea una tavola json

for i in oggetti:
    add_items_file_json(i)
    

# search_product_json() #metodo che serve per cercare un prodotto nel file json

# Viene creato il file nel caso in cui questo non dovesse esistere nel path
# Se il file esiste viene solo letto e ne viene fatto un update

if (path.isfile(filename) == False):    
    create_file(oggetti)
else:
    list_items = _read_file()
    # print(list_items)
    update_file(oggetti)
    
reset_file_json()






